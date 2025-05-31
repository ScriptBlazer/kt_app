from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from shuttle.models import Shuttle
from hotels.models import HotelBooking
from django.shortcuts import redirect
from jobs.models import Job
from django.utils import timezone
from shuttle.forms import DriverAssignmentForm
import pytz
import csv
from people.models import Driver
import datetime
from django.http import HttpResponse
import openpyxl
from openpyxl.utils import get_column_letter
from shuttle.models import Shuttle
from openpyxl.styles import Font, PatternFill

@login_required
def admin_page(request):
    # Set timezone to Budapest
    budapest_tz = pytz.timezone('Europe/Budapest')
    now_in_budapest = timezone.now().astimezone(budapest_tz)
    today = now_in_budapest.date()

    # Fetch today's jobs
    driving_jobs_today = Job.objects.filter(job_date=today, is_confirmed=True)

    context = {
        'driving_jobs_today': driving_jobs_today,
    }

    return render(request, 'admin.html', context)


@login_required
def services_page(request):

     # Set timezone to Budapest
    budapest_tz = pytz.timezone('Europe/Budapest')
    now_in_budapest = timezone.now().astimezone(budapest_tz)
    today = now_in_budapest.date()

    # Fetch today's shuttle jobs
    shuttle_jobs_today = Shuttle.objects.filter(shuttle_date=today, is_confirmed=True)
    
    # Fetch today's hotel bookings
    hotel_bookings_today = HotelBooking.objects.filter(check_in__date=today, is_confirmed=True)

    context = {
        'shuttle_jobs_today': shuttle_jobs_today,
        'hotel_bookings_today': hotel_bookings_today,
    }

    return render(request, 'services.html', context)


# Export_page view
@login_required
def export_jobs(request):
    month_range = [(0, 'All')] + [(i, datetime.date(1900, i, 1).strftime('%B')) for i in range(1, 13)]
    now = timezone.now()
    current_year = now.year
    current_month = now.month

    # If it's a GET form render, return the export page
    if request.method == 'GET' and not request.GET.get('format'):
        return render(request, 'export_page.html', {
            'month_range': month_range,
            'current_year': current_year,
            'current_month': current_month,
        })

    filter_type = request.GET.get('filter', 'all')
    file_format = request.GET.get('format', 'csv')
    year = request.GET.get('year')
    month = request.GET.get('month')

    if not year or not file_format:
        if not year and not month:
            error_message = 'Please select a year and month before exporting.'
        elif not year:
            error_message = 'Please select a year before exporting.'
        elif not month:
            error_message = 'Please select a month before exporting.'
        else:
            error_message = 'Missing required export information.'

        return render(request, 'export_page.html', {
            'month_range': month_range,
            'current_year': current_year,
            'current_month': current_month,
            'error_message': error_message
        })

    if file_format == 'csv':
        jobs = Job.objects.filter(is_confirmed=True).order_by('-job_date', '-job_time')
    else:
        jobs = Job.objects.filter(is_confirmed=True).order_by('-job_date', '-job_time')

    try:
        year_int = int(year)
        jobs = jobs.filter(job_date__year=year_int).order_by('-job_date', '-job_time')
        if month and month != "0":
            month_int = int(month)
            jobs = jobs.filter(job_date__month=month_int)
            sheet_title = f"KT Driving Jobs {datetime.date(year_int, month_int, 1).strftime('%B %Y')}"
        else:
            sheet_title = f"KT Driving Jobs {year_int}"
    except ValueError:
        return HttpResponse("Invalid year or month", status=400)

    if file_format == 'csv':
        filename = sheet_title + ".csv"
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        writer = csv.writer(response)
        writer.writerow([
            'Customer Name', 'Customer Number', 'Job Date', 'Job Time',
            'No. of Passengers', 'Vehicle Type', 'Kilometers', 'Pickup',
            'Drop-off', 'Flight Number', 'Price', 'Currency', 'Payment Type',
            'Confirmed', 'Paid', 'Driver', 'Number Plate',
            'Agent Name', 'Agent Fee', 'Payments Summary'
        ])
        for job in jobs:
            payments = list(job.payments.all())
            payments_summary = ''
            if payments:
                payment_strs = []
                for idx, p in enumerate(payments, 1):
                    paid_to_name = (
                        p.paid_to_driver.name if p.paid_to_driver else
                        p.paid_to_agent.name if p.paid_to_agent else
                        p.paid_to_staff.name if p.paid_to_staff else
                        'Not specified'
                    )
                    amt = f"{p.payment_amount}" if p.payment_amount is not None else ''
                    curr = p.payment_currency if p.payment_currency else ''
                    paytype = p.payment_type if p.payment_type else ''
                    # Format: Payment 1: €50 (Card, to Driver: John)
                    payment_str = f"Payment {idx}: {curr}{amt} ({paytype}, to "
                    if p.paid_to_driver:
                        payment_str += f"Driver: {p.paid_to_driver.name}"
                    elif p.paid_to_agent:
                        payment_str += f"Agent: {p.paid_to_agent.name}"
                    elif p.paid_to_staff:
                        payment_str += f"Staff: {p.paid_to_staff.name}"
                    else:
                        payment_str += "Not specified"
                    payment_str += ")"
                    payment_strs.append(payment_str)
                payments_summary = ' | '.join(payment_strs)
            writer.writerow([
                job.customer_name,
                job.customer_number,
                job.job_date,
                job.job_time,
                job.no_of_passengers,
                job.vehicle_type,
                job.kilometers,
                job.pick_up_location,
                job.drop_off_location,
                job.flight_number,
                job.job_price,
                job.job_currency,
                job.payment_type,
                job.is_confirmed,
                job.is_paid,
                job.driver.name if job.driver else '',
                job.number_plate,
                job.agent_name.name if job.agent_name else '',
                f"{job.agent_percentage}%" if job.agent_percentage else '',
                payments_summary
            ])
        return response

    elif file_format == 'xlsx':
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_title[:31]
        headers = [
            'Customer Name', 'Customer Number', 'Job Date', 'Job Time',
            'No. of Passengers', 'Vehicle Type', 'Kilometers', 'Pickup',
            'Drop-off', 'Flight Number', 'Price', 'Currency', 'Payment Type',
            'Confirmed', 'Paid', 'Driver', 'Number Plate',
            'Agent Name', 'Agent Fee', 'Payments Summary'
        ]
        ws.append(headers)

        # Style header row
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")  # Orange
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill

        # Freeze top row
        ws.freeze_panes = "A2"

        # Apply auto-filter to the full header row (A1:R1) as openpyxl only supports a contiguous range.
        
        last_column = get_column_letter(len(headers))
        ws.auto_filter.ref = f"A1:{last_column}1"

        for job in jobs:
            payments = list(job.payments.all())
            payments_summary = ''
            if payments:
                payment_strs = []
                for idx, p in enumerate(payments, 1):
                    paid_to_name = (
                        p.paid_to_driver.name if p.paid_to_driver else
                        p.paid_to_agent.name if p.paid_to_agent else
                        p.paid_to_staff.name if p.paid_to_staff else
                        'Not specified'
                    )
                    amt = f"{p.payment_amount}" if p.payment_amount is not None else ''
                    curr = p.payment_currency if p.payment_currency else ''
                    paytype = p.payment_type if p.payment_type else ''
                    payment_str = f"Payment {idx}: {curr}{amt} ({paytype}, to "
                    if p.paid_to_driver:
                        payment_str += f"Driver: {p.paid_to_driver.name}"
                    elif p.paid_to_agent:
                        payment_str += f"Agent: {p.paid_to_agent.name}"
                    elif p.paid_to_staff:
                        payment_str += f"Staff: {p.paid_to_staff.name}"
                    else:
                        payment_str += "Not specified"
                    payment_str += ")"
                    payment_strs.append(payment_str)
                payments_summary = ' | '.join(payment_strs)
            ws.append([
                job.customer_name,
                job.customer_number,
                job.job_date,
                job.job_time.strftime('%H:%M'),
                job.no_of_passengers,
                job.vehicle_type,
                float(job.kilometers or 0),
                job.pick_up_location,
                job.drop_off_location,
                job.flight_number,
                float(job.job_price or 0),
                job.job_currency,
                job.payment_type,
                job.is_confirmed,
                job.is_paid,
                job.driver.name if job.driver else '',
                job.number_plate,
                job.agent_name.name if job.agent_name else '',
                f"{job.agent_percentage}%" if job.agent_percentage else '',
                payments_summary
            ])

        # Auto-size columns based on max content width
        for col_num, column_title in enumerate(headers, 1):
            column_letter = get_column_letter(col_num)
            max_length = len(column_title)
            for row in ws.iter_rows(min_col=col_num, max_col=col_num):
                for cell in row:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column_letter].width = max_length + 2

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        filename = sheet_title + ".xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response

    return HttpResponse("Invalid format", status=400)



@login_required
def export_shuttles(request):
    import csv
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse

    month_range = [(0, 'All')] + [(i, datetime.date(1900, i, 1).strftime('%B')) for i in range(1, 13)]
    now = timezone.now()
    current_year = now.year
    current_month = now.month

    # Filters from GET params
    file_format = request.GET.get('format', 'csv')
    year = request.GET.get('year')
    month = request.GET.get('month')
    filter_driver_id = request.GET.get('filter_driver')
    filter_direction = request.GET.get('filter_direction')
    filter_is_paid = request.GET.get('filter_is_paid')

    if request.method == 'GET' and not request.GET.get('format'):
        drivers = Driver.objects.order_by('name')
        return render(request, 'export_page.html', {
            'month_range': month_range,
            'current_year': current_year,
            'current_month': current_month,
            'drivers': drivers,
        })

    if not year:
        return render(request, 'export_page.html', {
            'month_range': month_range,
            'current_year': current_year,
            'current_month': current_month,
            'error_message': 'Please select a year for export.'
        })

    shuttles = Shuttle.objects.filter(is_confirmed=True).order_by('-shuttle_date')

    try:
        year_int = int(year)
        shuttles = shuttles.filter(shuttle_date__year=year_int)
        if month and month != "0":
            month_int = int(month)
            shuttles = shuttles.filter(shuttle_date__month=month_int)
            sheet_title = f"KT Shuttles {datetime.date(year_int, month_int, 1).strftime('%B %Y')}"
        else:
            sheet_title = f"KT Shuttles {year_int}"
    except ValueError:
        return HttpResponse("Invalid year or month", status=400)

    if filter_driver_id:
        shuttles = shuttles.filter(driver_id=filter_driver_id)
    if filter_direction:
        shuttles = shuttles.filter(shuttle_direction=filter_direction)
    if filter_is_paid:
        if filter_is_paid == '1':
            shuttles = shuttles.filter(is_paid=True)
        elif filter_is_paid == '0':
            shuttles = shuttles.filter(is_paid=False)

    if file_format == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{sheet_title}.csv"'
        writer = csv.writer(response)
        writer.writerow(['Customer Name', 'Customer Number', 'Date', 'Direction', 'No. of Passengers', 'Price', 'Currency', 'Confirmed', 'Paid', 'Completed', 'Driver', 'Payments Summary'])

        for shuttle in shuttles:
            payments = list(shuttle.payments.all())
            payments_summary = ''
            if payments:
                payment_strs = []
                for idx, p in enumerate(payments, 1):
                    paid_to_name = (
                        p.paid_to_driver.name if p.paid_to_driver else
                        p.paid_to_agent.name if p.paid_to_agent else
                        p.paid_to_staff.name if p.paid_to_staff else
                        'Not specified'
                    )
                    amt = f"{p.payment_amount}" if p.payment_amount is not None else ''
                    curr = p.payment_currency if p.payment_currency else ''
                    paytype = p.payment_type if p.payment_type else ''
                    payment_str = f"Payment {idx}: {curr}{amt} ({paytype}, to "
                    if p.paid_to_driver:
                        payment_str += f"Driver: {p.paid_to_driver.name}"
                    elif p.paid_to_agent:
                        payment_str += f"Agent: {p.paid_to_agent.name}"
                    elif p.paid_to_staff:
                        payment_str += f"Staff: {p.paid_to_staff.name}"
                    else:
                        payment_str += "Not specified"
                    payment_str += ")"
                    payment_strs.append(payment_str)
                payments_summary = ' | '.join(payment_strs)

            writer.writerow([
                shuttle.customer_name,
                shuttle.customer_number,
                shuttle.shuttle_date,
                shuttle.shuttle_direction,
                shuttle.no_of_passengers,
                shuttle.price,
                shuttle.price_currency if hasattr(shuttle, 'price_currency') else 'EUR',
                shuttle.is_confirmed,
                shuttle.is_paid,
                shuttle.is_completed,
                shuttle.driver.name if shuttle.driver else '',
                payments_summary
            ])
        return response

    elif file_format == 'xlsx':
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_title[:31]
        headers = ['Customer Name', 'Customer Number', 'Date', 'Direction', 'No. of Passengers', 'Price', 'Currency', 'Confirmed', 'Paid', 'Completed', 'Driver', 'Payments Summary']
        ws.append(headers)

        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
        ws.freeze_panes = "A2"
        last_column = get_column_letter(len(headers))
        ws.auto_filter.ref = f"A1:{last_column}1"

        for shuttle in shuttles:
            payments = list(shuttle.payments.all())
            payments_summary = ''
            if payments:
                payment_strs = []
                for idx, p in enumerate(payments, 1):
                    amt = f"{p.payment_amount}" if p.payment_amount is not None else ''
                    curr = p.payment_currency if p.payment_currency else ''
                    paytype = p.payment_type if p.payment_type else ''
                    payment_str = f"Payment {idx}: {curr}{amt} ({paytype}, to "
                    if p.paid_to_driver:
                        payment_str += f"Driver: {p.paid_to_driver.name}"
                    elif p.paid_to_agent:
                        payment_str += f"Agent: {p.paid_to_agent.name}"
                    elif p.paid_to_staff:
                        payment_str += f"Staff: {p.paid_to_staff.name}"
                    else:
                        payment_str += "Not specified"
                    payment_str += ")"
                    payment_strs.append(payment_str)
                payments_summary = ' | '.join(payment_strs)

            ws.append([
                shuttle.customer_name,
                shuttle.customer_number,
                shuttle.shuttle_date,
                shuttle.shuttle_direction,
                shuttle.no_of_passengers,
                shuttle.price,
                shuttle.price_currency if hasattr(shuttle, 'price_currency') else 'EUR',
                shuttle.is_confirmed,
                shuttle.is_paid,
                shuttle.is_completed,
                shuttle.driver.name if shuttle.driver else '',
                payments_summary
            ])

        for col_num, column_title in enumerate(headers, 1):
            column_letter = get_column_letter(col_num)
            max_length = len(column_title)
            for row in ws.iter_rows(min_col=col_num, max_col=col_num):
                for cell in row:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column_letter].width = max_length + 2

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{sheet_title}.xlsx"'
        wb.save(response)
        return response

    return HttpResponse("Invalid format", status=400)